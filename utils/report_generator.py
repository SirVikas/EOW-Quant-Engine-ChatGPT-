"""
EOW Quant Engine — Triple-Format Report Generator
Produces XLSX + PDF + Markdown bundled in a single ZIP archive.
No external PDF library required — uses a pure-Python minimal PDF writer.
"""
from __future__ import annotations

import io
import math
import time
import zipfile
from typing import List


# ─────────────────────────────────────────────────────────────────────────────
# Minimal Pure-Python PDF Writer
# Generates a valid PDF 1.4 document with Helvetica text, no images.
# ─────────────────────────────────────────────────────────────────────────────

class _PDFWriter:
    """
    Produces a minimal but fully spec-compliant PDF 1.4 file using only
    Python stdlib.  Supports:
      • Multiple pages
      • Helvetica / Helvetica-Bold (built-in Type 1 fonts, no embedding needed)
      • Line drawing (thin rules)
    """

    PAGE_W  = 595   # A4 width  (points)
    PAGE_H  = 842   # A4 height (points)
    MARGIN  = 60

    def __init__(self):
        self._pages:   List[str] = []   # page stream text per page
        self._cur:     list      = []   # current page lines
        self._y:       float     = self.PAGE_H - self.MARGIN
        self._objects: List[bytes] = []
        self._offsets: List[int]   = []

    # ── High-level helpers ────────────────────────────────────────────────────

    def title(self, text: str):
        self._text(self.MARGIN, self._y, text, size=18, bold=True)
        self._y -= 26

    def h2(self, text: str):
        self._y -= 6
        self._line_rule()
        self._text(self.MARGIN, self._y, text, size=13, bold=True)
        self._y -= 20

    def body(self, text: str, size: int = 10):
        # Wrap long lines
        words = text.split()
        line, max_w = [], 95
        for w in words:
            if len(' '.join(line + [w])) > max_w:
                self._text(self.MARGIN, self._y, ' '.join(line), size=size)
                self._y -= size + 3
                line = [w]
            else:
                line.append(w)
        if line:
            self._text(self.MARGIN, self._y, ' '.join(line), size=size)
            self._y -= size + 3
        self._check_newpage()

    def kv(self, label: str, value: str, size: int = 10):
        row = f"{label:<32}{value}"
        self._text(self.MARGIN, self._y, row, size=size)
        self._y -= size + 3
        self._check_newpage()

    def blank(self, n: int = 1):
        self._y -= 12 * n
        self._check_newpage()

    def table_header(self, cols: List[str], widths: List[int]):
        self._table_row(cols, widths, bold=True, fill=True)

    def table_row(self, cols: List[str], widths: List[int]):
        self._table_row(cols, widths, bold=False, fill=False)

    def new_page(self):
        self._flush_page()
        self._y = self.PAGE_H - self.MARGIN

    def build(self) -> bytes:
        self._flush_page()
        buf = io.BytesIO()

        # Header
        buf.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")

        n_pages = len(self._pages)

        # Object 1 — Catalog
        self._offsets.append(buf.tell())
        buf.write(b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n")

        # Object 2 — Pages dict (placeholder; we write it last)
        pages_pos = buf.tell()
        self._offsets.append(pages_pos)
        kids = " ".join(f"{3 + i} 0 R" for i in range(n_pages))
        pages_dict = (
            f"2 0 obj\n<< /Type /Pages /Kids [{kids}] /Count {n_pages} >>\nendobj\n"
        ).encode()
        buf.write(pages_dict)

        # Font object index (placed after all page objects)
        font_obj_num = 3 + n_pages

        # Objects 3…(3+n_pages-1) — Page objects, each referencing stream
        for i, stream_text in enumerate(self._pages):
            page_obj   = 3 + i
            stream_obj = font_obj_num + 1 + i   # stream object follows all page objs

            self._offsets.append(buf.tell())
            page = (
                f"{page_obj} 0 obj\n"
                f"<< /Type /Page /Parent 2 0 R "
                f"/MediaBox [0 0 {self.PAGE_W} {self.PAGE_H}] "
                f"/Contents {stream_obj} 0 R "
                f"/Resources << /Font << "
                f"/F1 {font_obj_num} 0 R "
                f"/F2 {font_obj_num} 0 R "
                f">> >> >>\nendobj\n"
            ).encode()
            buf.write(page)

        # Font object
        self._offsets.append(buf.tell())
        buf.write((
            f"{font_obj_num} 0 obj\n"
            "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica "
            "/Encoding /WinAnsiEncoding >>\nendobj\n"
        ).encode())

        # Stream objects for each page
        for i, stream_text in enumerate(self._pages):
            stream_obj = font_obj_num + 1 + i
            # Replace bold font calls — we encode bold separately below
            raw = stream_text.encode("latin-1", errors="replace")
            self._offsets.append(buf.tell())
            buf.write((
                f"{stream_obj} 0 obj\n"
                f"<< /Length {len(raw)} >>\nstream\n"
            ).encode())
            buf.write(raw)
            buf.write(b"\nendstream\nendobj\n")

        # xref table
        xref_pos = buf.tell()
        n_objs = len(self._offsets) + 1   # +1 for obj 0
        buf.write(f"xref\n0 {n_objs}\n".encode())
        buf.write(b"0000000000 65535 f \n")
        for off in self._offsets:
            buf.write(f"{off:010d} 00000 n \n".encode())

        buf.write((
            f"trailer\n<< /Size {n_objs} /Root 1 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF\n"
        ).encode())

        return buf.getvalue()

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _text(self, x: float, y: float, text: str, size: int = 10, bold: bool = False):
        # Escape PDF special chars
        text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        font = "/F1"   # Helvetica (bold uses same font — true bold needs embedding)
        if bold:
            # Simulate bold with slightly wider word spacing
            self._cur.append(
                f"BT {font} {size} Tf {x:.1f} {y:.1f} Td 0.3 Tw ({text}) Tj 0 Tw ET"
            )
        else:
            self._cur.append(f"BT {font} {size} Tf {x:.1f} {y:.1f} Td ({text}) Tj ET")

    def _line_rule(self):
        y = self._y + 2
        self._cur.append(
            f"{self.MARGIN} {y:.1f} m {self.PAGE_W - self.MARGIN} {y:.1f} l S"
        )

    def _table_row(self, cols: List[str], widths: List[int], bold: bool, fill: bool):
        x = self.MARGIN
        if fill:
            row_h = 14
            self._cur.append(
                f"0.92 0.92 0.92 rg "
                f"{x} {self._y - 3} {self.PAGE_W - 2*self.MARGIN} {row_h} re f "
                f"0 0 0 rg"
            )
        for col, w in zip(cols, widths):
            col = str(col)
            if len(col) > w // 6:
                col = col[:w // 6 - 1] + "…"
            self._text(x + 2, self._y, col, size=9, bold=bold)
            x += w
        self._y -= 14
        self._check_newpage()

    def _check_newpage(self):
        if self._y < self.MARGIN + 40:
            self.new_page()

    def _flush_page(self):
        if self._cur:
            self._pages.append("\n".join(self._cur))
            self._cur = []


# ─────────────────────────────────────────────────────────────────────────────
# XLSX Generator
# ─────────────────────────────────────────────────────────────────────────────

def _generate_xlsx(trades: list, stats: dict, thoughts: list) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: return a UTF-8 CSV wrapped in bytes
        return _generate_csv_fallback(trades).encode("utf-8")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Trade History ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Trade History"

    header_fill = PatternFill("solid", fgColor="B2F5EA")
    bold        = Font(bold=True)

    trade_cols = [
        "trade_id", "symbol", "side", "order_type", "regime",
        "entry_price", "exit_price", "qty",
        "gross_pnl", "fee_entry", "fee_exit", "slippage_cost", "borrow_cost", "net_pnl",
        "net_pnl_pct", "r_multiple", "strategy_id", "entry_ts", "exit_ts",
    ]
    for ci, col in enumerate(trade_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, t in enumerate(trades, 2):
        for ci, col in enumerate(trade_cols, 1):
            val = t.get(col, "")
            ws.cell(row=ri, column=ci, value=val)

    for i, _ in enumerate(trade_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # ── Sheet 2: Session Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Session Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    ws2["A1"], ws2["B1"] = "Metric", "Value"
    ws2["A1"].font = bold
    ws2["B1"].font = bold
    ws2["A1"].fill = header_fill
    ws2["B1"].fill = header_fill

    summary_rows = [
        ("Total Trades",       stats.get("total_trades", 0)),
        ("Win Rate (%)",       round(stats.get("win_rate", 0), 2)),
        ("Profit Factor",      round(stats.get("profit_factor", 0), 3)),
        ("Sharpe Ratio",       round(stats.get("sharpe_ratio", 0), 3)),
        ("Net PnL (USDT)",     round(stats.get("total_net_pnl", 0), 4)),
        ("Total Fees (USDT)",  round(stats.get("total_fees_paid", 0), 4)),
        ("Total Slippage (USDT)", round(stats.get("total_slippage", 0), 4)),
        ("Max Drawdown (%)",   round(stats.get("max_drawdown_pct", 0), 2)),
        ("Avg Win (USDT)",     round(stats.get("avg_win_usdt", 0), 4)),
        ("Avg Loss (USDT)",    round(stats.get("avg_loss_usdt", 0), 4)),
        ("Final Capital (USDT)", round(stats.get("capital", 0), 2)),
    ]
    for ri, (label, val) in enumerate(summary_rows, 2):
        ws2.cell(row=ri, column=1, value=label)
        ws2.cell(row=ri, column=2, value=val)

    # ── Sheet 3: Signal Audit Log ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Signal Audit")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 90

    ws3["A1"].value = "Timestamp"
    ws3["B1"].value = "Level"
    ws3["C1"].value = "Message"
    for cell in [ws3["A1"], ws3["B1"], ws3["C1"]]:
        cell.font = bold
        cell.fill = header_fill

    level_colours = {
        "TRADE":  "C6F6D5", "SIGNAL": "BEE3F8",
        "FILTER": "FEFCBF", "HALT":   "FED7D7",
        "SYSTEM": "E9D8FD",
    }
    for ri, t in enumerate(thoughts, 2):
        ts_s = time.strftime("%H:%M:%S", time.gmtime(t.get("ts", 0) / 1000))
        lvl  = t.get("level", "INFO")
        ws3.cell(row=ri, column=1, value=ts_s)
        c2 = ws3.cell(row=ri, column=2, value=lvl)
        ws3.cell(row=ri, column=3, value=t.get("msg", ""))
        colour = level_colours.get(lvl)
        if colour:
            c2.fill = PatternFill("solid", fgColor=colour)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_csv_fallback(trades: list) -> str:
    """CSV fallback when openpyxl is unavailable."""
    cols = ["trade_id", "symbol", "side", "net_pnl", "r_multiple", "entry_ts"]
    lines = [",".join(cols)]
    for t in trades:
        lines.append(",".join(str(t.get(c, "")) for c in cols))
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# PDF Generator
# ─────────────────────────────────────────────────────────────────────────────

def _generate_pdf(
    stats:     dict,
    mode_info: dict,
    analytics: dict,
    trades:    list,
) -> bytes:
    now_str = time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())
    pdf     = _PDFWriter()

    # ── Page 1: Executive Summary ─────────────────────────────────────────────
    pdf.title("EOW Quant Engine — Performance Report")
    pdf.body(f"Generated: {now_str}   |   Mode: {mode_info.get('label', '—')}", size=9)
    persist = mode_info.get("persistence_status", "")
    if persist:
        pdf.body(persist, size=9)
    pdf.blank()

    pdf.h2("1. Executive Summary")
    total_net  = stats.get("total_net_pnl", 0.0)
    direction  = "PROFIT" if total_net >= 0 else "LOSS"
    pdf.body(
        f"The engine closed {stats.get('total_trades', 0)} trades with a net {direction} "
        f"of {total_net:+.2f} USDT. Win rate: {stats.get('win_rate', 0):.1f}% | "
        f"Profit factor: {stats.get('profit_factor', 0):.2f} | "
        f"Sharpe: {stats.get('sharpe_ratio', 0):.3f} | "
        f"Max drawdown: {stats.get('max_drawdown_pct', 0):.2f}%."
    )
    pdf.blank()

    pdf.h2("2. Key Performance Metrics")
    kv_rows = [
        ("Final Capital (USDT)",   f"{stats.get('capital', 0):,.2f}"),
        ("Net PnL (USDT)",         f"{total_net:+,.4f}"),
        ("Total Trades",           str(stats.get("total_trades", 0))),
        ("Win Rate",               f"{stats.get('win_rate', 0):.1f}%"),
        ("Profit Factor",          f"{stats.get('profit_factor', 0):.3f}"),
        ("Sharpe Ratio",           f"{stats.get('sharpe_ratio', 0):.3f}"),
        ("Sortino Ratio",          f"{analytics.get('sortino_ratio', 0):.3f}"),
        ("Calmar Ratio",           f"{analytics.get('calmar_ratio', 0):.3f}"),
        ("Max Drawdown",           f"{stats.get('max_drawdown_pct', 0):.2f}%"),
        ("Risk of Ruin",           f"{analytics.get('risk_of_ruin_pct', 0):.2f}%"),
        ("Avg Win (USDT)",         f"{stats.get('avg_win_usdt', 0):+.4f}"),
        ("Avg Loss (USDT)",        f"{stats.get('avg_loss_usdt', 0):+.4f}"),
        ("Total Fees Paid (USDT)", f"{stats.get('total_fees_paid', 0):.4f}"),
        ("Total Slippage (USDT)",  f"{stats.get('total_slippage', 0):.4f}"),
        ("Deployability Index",    f"{analytics.get('deployability', {}).get('score', 0)}/100 "
                                   f"({analytics.get('deployability', {}).get('tier', '—')})"),
    ]
    for label, val in kv_rows:
        pdf.kv(label, val)

    # ── Page 2: Trade Log ─────────────────────────────────────────────────────
    pdf.new_page()
    pdf.h2("3. Trade History (last 50 trades)")

    cols   = ["Symbol", "Side", "Net PnL", "R-Multiple", "Regime", "Order Type"]
    widths = [80, 55, 80, 80, 100, 85]
    pdf.table_header(cols, widths)

    for t in trades[-50:]:
        pdf.table_row([
            t.get("symbol", ""),
            t.get("side", ""),
            f"{t.get('net_pnl', 0):+.2f}",
            f"{t.get('r_multiple', 0):.3f}",
            t.get("regime", ""),
            t.get("order_type", ""),
        ], widths)

    # ── Benchmark section ─────────────────────────────────────────────────────
    pdf.blank()
    pdf.h2("4. Benchmark Comparison")
    bm_data = analytics.get("benchmark", {})
    eng     = bm_data.get("engine", {})
    pdf.kv("EOW Engine Annual Return",  f"{eng.get('annual_return_pct', 0):+.1f}%")
    pdf.kv("EOW Engine Sharpe",         f"{eng.get('sharpe', 0):.3f}")
    pdf.kv("EOW Engine Max DD",         f"{eng.get('max_dd_pct', 0):.2f}%")
    pdf.blank()
    for bm in bm_data.get("benchmarks", []):
        pdf.kv(bm["name"], f"Return {bm['annual_return_pct']:+.1f}%  Sharpe {bm['sharpe']:.2f}")

    return pdf.build()


# ─────────────────────────────────────────────────────────────────────────────
# Markdown Generator
# ─────────────────────────────────────────────────────────────────────────────

def _generate_markdown(
    stats:     dict,
    mode_info: dict,
    analytics: dict,
    trades:    list,
    thoughts:  list,
) -> str:
    now_str = time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())
    lines   = []

    lines += [
        f"# EOW Quant Engine — Performance Report",
        f"",
        f"**Generated:** {now_str}  ",
        f"**Mode:** `{mode_info.get('label', '—')}`  ",
        f"**Persistence:** {mode_info.get('persistence_status', '—')}  ",
        f"",
        f"---",
        f"",
        f"## 1. Executive Summary",
        f"",
    ]

    total_net = stats.get("total_net_pnl", 0.0)
    direction = "PROFIT" if total_net >= 0 else "LOSS"
    lines.append(
        f"The engine closed **{stats.get('total_trades', 0)} trades** with a net "
        f"**{direction}** of **{total_net:+.2f} USDT**.  "
    )
    lines += [
        f"",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Final Capital | ${stats.get('capital', 0):,.2f} USDT |",
        f"| Net PnL | {total_net:+,.4f} USDT |",
        f"| Win Rate | {stats.get('win_rate', 0):.1f}% |",
        f"| Profit Factor | {stats.get('profit_factor', 0):.3f} |",
        f"| Sharpe Ratio | {stats.get('sharpe_ratio', 0):.3f} |",
        f"| Sortino Ratio | {analytics.get('sortino_ratio', 0):.3f} |",
        f"| Calmar Ratio | {analytics.get('calmar_ratio', 0):.3f} |",
        f"| Max Drawdown | {stats.get('max_drawdown_pct', 0):.2f}% |",
        f"| Risk of Ruin | {analytics.get('risk_of_ruin_pct', 0):.2f}% |",
        f"| Total Fees | {stats.get('total_fees_paid', 0):.4f} USDT |",
        f"| Total Slippage | {stats.get('total_slippage', 0):.4f} USDT |",
        f"| Deployability | {analytics.get('deployability', {}).get('score', 0)}/100 "
        f"({analytics.get('deployability', {}).get('tier', '—')}) |",
        f"",
        f"---",
        f"",
        f"## 2. Performance Audit",
        f"",
        f"### 2.1 PnL Breakdown",
        f"",
        f"- **Gross PnL:** {total_net + stats.get('total_fees_paid', 0) + stats.get('total_slippage', 0):+.4f} USDT (before all costs)",
        f"- **Fees deducted:** -{stats.get('total_fees_paid', 0):.4f} USDT",
        f"- **Slippage deducted:** -{stats.get('total_slippage', 0):.4f} USDT",
        f"- **Net PnL (bankable):** {total_net:+.4f} USDT",
        f"",
        f"### 2.2 Trade Statistics",
        f"",
        f"- Avg win: {stats.get('avg_win_usdt', 0):+.4f} USDT",
        f"- Avg loss: {stats.get('avg_loss_usdt', 0):+.4f} USDT",
        f"- Profit factor: {stats.get('profit_factor', 0):.3f}",
        f"",
    ]

    # Benchmark section
    lines += [
        f"---",
        f"",
        f"## 3. Benchmark Comparison",
        f"",
        f"| Fund | Annual Return | Sharpe | Sortino | Max DD |",
        f"|------|--------------|--------|---------|--------|",
    ]
    bm_data = analytics.get("benchmark", {})
    eng = bm_data.get("engine", {})
    lines.append(
        f"| **EOW Engine** | **{eng.get('annual_return_pct', 0):+.1f}%** | "
        f"**{eng.get('sharpe', 0):.2f}** | **{eng.get('sortino', 0):.2f}** | "
        f"**{eng.get('max_dd_pct', 0):.1f}%** |"
    )
    for bm in bm_data.get("benchmarks", []):
        lines.append(
            f"| {bm['name']} | {bm['annual_return_pct']:+.1f}% | "
            f"{bm['sharpe']:.2f} | {bm['sortino']:.2f} | {bm['max_dd_pct']:.1f}% |"
        )

    # Signal audit
    lines += [
        f"",
        f"---",
        f"",
        f"## 4. Signal Audit (CT-Scan Log)",
        f"",
        f"| Time | Level | Message |",
        f"|------|-------|---------|",
    ]
    for t in thoughts[-200:]:
        ts_s = time.strftime("%H:%M:%S", time.gmtime(t.get("ts", 0) / 1000))
        msg  = t.get("msg", "").replace("|", "\\|")
        lvl  = t.get("level", "INFO")
        lines.append(f"| {ts_s} | {lvl} | {msg} |")

    lines += [
        f"",
        f"---",
        f"*EOW Quant Engine V4.0 — {now_str}*",
    ]

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# Public ZIP Builder
# ─────────────────────────────────────────────────────────────────────────────

def build_report_archive(
    trades:    list,
    stats:     dict,
    mode_info: dict,
    analytics: dict,
    thoughts:  list,
) -> bytes:
    """
    Return a ZIP archive (bytes) containing:
      • eow_trades.xlsx       — full trade history, summary, signal audit
      • eow_report.pdf        — executive summary with metrics
      • eow_report.md         — markdown developer log
    """
    ts_tag = time.strftime("%Y%m%d_%H%M%S", time.gmtime())

    xlsx_bytes = _generate_xlsx(trades, stats, thoughts)
    pdf_bytes  = _generate_pdf(stats, mode_info, analytics, trades)
    md_text    = _generate_markdown(stats, mode_info, analytics, trades, thoughts)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"eow_trades_{ts_tag}.xlsx", xlsx_bytes)
        zf.writestr(f"eow_report_{ts_tag}.pdf",  pdf_bytes)
        zf.writestr(f"eow_report_{ts_tag}.md",   md_text.encode("utf-8"))

    return buf.getvalue()
